#importações necessárias
from openpyxl import *
from requisição import *

#===============================================================================================================================

#Função para inserir dados na planilha

#função para setagem de itens

def set (planilha, coluna, nome, url, linhaa =2):
 #setagem de variáveis e listas
 linha = linhaa
 lista = []
 
 #1º linha(Cabeçalho)
 planilha.cell(row=1, column=coluna, value = nome)
 #inserção na lista que será usada depois
 lista.insert(0, nome)

 #setagem de itens de acordo com a coluna
 for user in url:
  #pegar o tipo de dado fornecido
  a = str(user[nome])
  #setagem dos dados
  planilha.cell(row=linha, column=coluna, value=a)
  #atualização para indicar próximo campo
  linha +=1
  #adicionamento dos dados em uma lista para comparação posterior
  lista.insert(0, a)
 
 #descobrimento do maior item da coluna
 maximo = list(map(lambda lista: len(lista), lista))
 maximo = max(maximo)
 #função de suporte para definição do tamanho de acordo com o maior dado
 ajuste(planilha, maximo, coluna)

#função de suporte a função principal
def ajuste (planilha, a, b):
 if b ==1:
  b = "A"
 if b ==2:
  b = "B"
 if b ==3:
  b = "C"
 if b ==4:
  b = "D"
 if b ==5:
  b = "E"
 if b == 6:
  b = "F"
 planilha.column_dimensions[b].width = a+3
 #+3 é para deixar adequado o suficiente para leitura, visto que alguns dados podem ser confundidos

#============================================================================================================================

#planilha:

#criação de planilha
book = Workbook()

#ativação para alteração
planilha = book.active
                            
#inserção de dados da primeira url
set(planilha, coluna=1, nome = "id", url = url1)
set(planilha, coluna=2, nome = "email", url = url1 )
set(planilha, coluna=3, nome = "first_name", url = url1)
set(planilha, coluna=4, nome = "last_name", url = url1)
set(planilha, coluna=5, nome = "avatar", url = url1)

#inserção de dados da segunda url
set(planilha, coluna=1, nome = "id", url = url2, linhaa=8)
set(planilha, coluna=2, nome = "email", url = url2, linhaa=8)
set(planilha, coluna=3, nome = "first_name", url = url2, linhaa=8)
set(planilha, coluna=4, nome = "last_name", url = url2, linhaa=8)
set(planilha, coluna=5, nome = "avatar", url = url2, linhaa=8)

#salvar documento
book.save("Planilha.xlsx")